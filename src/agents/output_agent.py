"""
src/agents/output_agent.py - Results writer
Merges AI analysis results back into the original DataFrame and writes Excel.
"""
from __future__ import annotations
import json
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.utils.config import OUTPUT_DIR
from src.utils.cache_manager import CacheManager


# Kolom AI yang akan ditambahkan ke Excel
AI_COLUMNS = [
    "AI_Tipe_Masalah",
    "AI_Kategori_Bug",
    "AI_Sub_Kategori",
    "AI_Root_Cause",
    "AI_Summary",
    "AI_Urgensi",
    "AI_Tags",
    "AI_Aplikasi_Terkait",
    "AI_Processed_At",
]

URGENCY_COLORS = {
    "Critical": "C00000",  # Dark Red
    "High": "FF0000",      # Red
    "Medium": "FF8C00",    # Orange
    "Low": "70AD47",       # Green
}

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
AI_HEADER_FILL = PatternFill(start_color="375623", end_color="375623", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)


def merge_results(df: pd.DataFrame, cache: CacheManager) -> pd.DataFrame:
    """Merge AI analysis results from cache back into the DataFrame."""
    results = cache.get_all_results()
    if not results:
        print("⚠️  Tidak ada hasil analisis di cache")
        return df

    result_map = {r["no_tiket"]: r for r in results}

    def map_col(no_tiket, field, default=""):
        r = result_map.get(str(no_tiket), {})
        val = r.get(field, default)
        if isinstance(val, list):
            return ", ".join(str(v) for v in val)
        return str(val) if val else default

    df = df.copy()
    df["AI_Tipe_Masalah"]   = df["no_tiket"].apply(lambda x: map_col(x, "tipe_masalah"))
    df["AI_Kategori_Bug"]   = df["no_tiket"].apply(lambda x: map_col(x, "kategori_bug", "-"))
    df["AI_Sub_Kategori"]   = df["no_tiket"].apply(lambda x: map_col(x, "sub_kategori"))
    df["AI_Root_Cause"]     = df["no_tiket"].apply(lambda x: map_col(x, "root_cause"))
    df["AI_Summary"]        = df["no_tiket"].apply(lambda x: map_col(x, "summary"))
    df["AI_Urgensi"]        = df["no_tiket"].apply(lambda x: map_col(x, "urgensi", "Low"))
    df["AI_Tags"]           = df["no_tiket"].apply(lambda x: map_col(x, "tags"))
    df["AI_Aplikasi_Terkait"] = df["no_tiket"].apply(lambda x: map_col(x, "aplikasi_terkait", "N/A"))
    df["AI_Processed_At"]  = df["no_tiket"].apply(
        lambda x: result_map.get(str(x), {}).get("processed_at", "")
    )

    return df


def write_excel(df: pd.DataFrame, output_path: Path | None = None) -> Path:
    """Write the merged DataFrame to a styled Excel file."""
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        output_path = OUTPUT_DIR / f"Incident_Analysis_{timestamp}.xlsx"

    print(f"\n📝 Menulis Excel → {output_path.name}...")

    # Build final column list
    # Original columns we want to keep
    original_cols = [c for c in df.columns if not c.startswith("AI_") and c not in
                     ["deskripsi_raw", "resolved_raw", "deskripsi", "resolved_notes"]]
    ai_cols = [c for c in AI_COLUMNS if c in df.columns]

    # Add cleaned text columns for readability
    readable_cols = []
    if "deskripsi" in df.columns:
        readable_cols.append("deskripsi")
    if "resolved_notes" in df.columns:
        readable_cols.append("resolved_notes")

    final_cols = original_cols + readable_cols + ai_cols
    final_cols = [c for c in final_cols if c in df.columns]  # safety check

    df_out = df[final_cols].copy()

    # Write with openpyxl for styling
    df_out.to_excel(output_path, index=False, sheet_name="Analisis Tiket")

    # Apply styling
    _style_excel(str(output_path), len(original_cols) + len(readable_cols), len(ai_cols))

    print(f"   ✅ {len(df_out):,} baris ditulis ke {output_path}")
    return output_path


def _style_excel(path: str, n_original_cols: int, n_ai_cols: int):
    """Apply professional styling to the output Excel."""
    wb = load_workbook(path)
    ws = wb.active

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, cell in enumerate(ws[1], start=1):
        is_ai_col = col_idx > n_original_cols
        cell.fill = AI_HEADER_FILL if is_ai_col else HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    # Freeze top row and set row height
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    # Auto-fit column widths (approximate)
    col_widths = {
        "no_tiket": 15, "judul": 35, "deskripsi": 45, "resolved_notes": 45,
        "status": 10, "kelompok": 25, "pelapor": 18, "tiket_dibuat": 18,
        "AI_Tipe_Masalah": 22, "AI_Kategori_Bug": 30, "AI_Sub_Kategori": 28,
        "AI_Root_Cause": 45, "AI_Summary": 45, "AI_Urgensi": 12,
        "AI_Tags": 25, "AI_Aplikasi_Terkait": 20,
    }

    for col_idx, cell in enumerate(ws[1], start=1):
        col_letter = get_column_letter(col_idx)
        header_name = cell.value or ""
        width = col_widths.get(header_name, 18)
        ws.column_dimensions[col_letter].width = width

    # Color urgency cells
    urgency_col = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "AI_Urgensi":
            urgency_col = col_idx
            break

    if urgency_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                 min_col=urgency_col, max_col=urgency_col):
            for cell in row:
                val = str(cell.value or "")
                color = URGENCY_COLORS.get(val, "FFFFFF")
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF" if val in ["Critical", "High"] else "000000")

    # Style data rows (alternating background subtly)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        bg = "F8F9FA" if row_idx % 2 == 0 else "FFFFFF"
        for cell in row:
            cell.border = border
            if cell.fill.fill_type == "none" or not cell.fill.start_color.rgb:
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    wb.save(path)


def write_json_for_dashboard(df: pd.DataFrame, cache: CacheManager) -> Path:
    """Write analysis results as JSON for the web dashboard to consume."""
    results = cache.get_all_results()
    out_path = OUTPUT_DIR / "dashboard_data.json"

    # Build full records
    ticket_map = {}
    for _, row in df.iterrows():
        no_tiket = str(row.get("no_tiket", ""))
        ticket_map[no_tiket] = {
            "no_tiket": no_tiket,
            "tiket_dibuat": str(row.get("tiket_dibuat", "")),
            "status": str(row.get("status", "")),
            "pelapor": str(row.get("pelapor", "")),
            "kelompok": str(row.get("kelompok", "")),
            "service": str(row.get("service", "")),
            "lokasi": str(row.get("lokasi", "")),
            "judul": str(row.get("judul", "")),
            "source_file": str(row.get("_source_file", "")),
        }

    full_records = []
    for r in results:
        no_tiket = r.get("no_tiket", "")
        base = ticket_map.get(no_tiket, {"no_tiket": no_tiket})
        merged = {**base, **r}
        # Ensure tipe_masalah field is present (backward compat)
        if not merged.get("tipe_masalah"):
            merged["tipe_masalah"] = merged.get("kategori_utama", "Bug Aplikasi")
        if not merged.get("kategori_bug"):
            merged["kategori_bug"] = "-"
        if isinstance(merged.get("tags"), list):
            merged["tags"] = merged["tags"]
        else:
            merged["tags"] = []
        full_records.append(merged)

    payload = {
        "generated_at": datetime.now().isoformat(),
        "total_tickets": len(full_records),
        "records": full_records,
    }

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, default=str)

    print(f"   ✅ Dashboard JSON ditulis → {out_path.name} ({len(full_records):,} records)")
    return out_path
